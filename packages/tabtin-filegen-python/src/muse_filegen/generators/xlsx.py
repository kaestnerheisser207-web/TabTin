"""XLSX generator (openpyxl)."""

from __future__ import annotations

from typing import Any, Dict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from muse_filegen.errors import SpecError
from muse_filegen.validate import as_text, require_list, require_mapping

_READ_HELP = """xlsx read 输出：
{
  "file_type": "xlsx",
  "sheets": [{"name": "Sheet1", "rows": [["甲", 1], ...]}]
}"""

_SPEC_HELP = """xlsx spec:
{
  "sheets": [
    {
      "name": "Sheet1",                 // optional sheet title
      "columns": [{"width": 20}, ...],  // optional per-column widths
      "header": ["Col A", "Col B"],     // optional, rendered bold
      "rows": [["a", 1], ["b", 2]]      // 2D array of cell values
    }
  ]
}"""


class XlsxGenerator:
    file_type = "xlsx"
    extensions = (".xlsx",)

    def spec_help(self) -> str:
        return _SPEC_HELP

    def generate(self, spec: Dict[str, Any], out_path: str) -> None:
        sheets = require_list(spec.get("sheets", []), "sheets")
        if not sheets:
            raise SpecError("xlsx spec needs at least one entry in `sheets`")

        workbook = Workbook()
        workbook.remove(workbook.active)

        for index, raw_sheet in enumerate(sheets):
            sheet_spec = require_mapping(raw_sheet, f"sheets[{index}]")
            title = str(sheet_spec.get("name") or f"Sheet{index + 1}")
            worksheet = workbook.create_sheet(title=_safe_title(title))

            header = sheet_spec.get("header")
            if header is not None:
                header_cells = require_list(header, f"sheets[{index}].header")
                worksheet.append([as_text(cell) for cell in header_cells])
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)

            rows = require_list(sheet_spec.get("rows", []), f"sheets[{index}].rows")
            for row_index, raw_row in enumerate(rows):
                row = require_list(raw_row, f"sheets[{index}].rows[{row_index}]")
                worksheet.append([_cell_value(cell) for cell in row])

            columns = sheet_spec.get("columns")
            if columns is not None:
                for col_index, raw_col in enumerate(
                    require_list(columns, f"sheets[{index}].columns")
                ):
                    col_spec = require_mapping(
                        raw_col, f"sheets[{index}].columns[{col_index}]"
                    )
                    width = col_spec.get("width")
                    if width is not None:
                        letter = get_column_letter(col_index + 1)
                        worksheet.column_dimensions[letter].width = float(width)

        workbook.save(out_path)


class XlsxReader:
    file_type = "xlsx"
    extensions = (".xlsx",)

    def read_help(self) -> str:
        return _READ_HELP

    def read(self, path: str) -> Dict[str, Any]:
        workbook = load_workbook(path, data_only=True, read_only=True)
        sheets = []
        for worksheet in workbook.worksheets:
            rows = [
                [_normalize_read_cell(cell) for cell in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
            sheets.append({"name": worksheet.title, "rows": rows})
        workbook.close()
        return {"file_type": "xlsx", "sheets": sheets}


def _normalize_read_cell(cell: Any) -> Any:
    """Keep JSON-serializable scalars; stringify the rest (dates, etc.)."""
    if cell is None or isinstance(cell, (str, int, float, bool)):
        return cell
    return str(cell)


def _cell_value(cell: Any) -> Any:
    """Preserve numbers/bools so Excel stores them as real types."""
    if isinstance(cell, (int, float, bool)) or cell is None:
        return cell
    return str(cell)


def _safe_title(title: str) -> str:
    """Excel sheet titles cap at 31 chars and forbid a few characters."""
    cleaned = title
    for ch in '[]:*?/\\':
        cleaned = cleaned.replace(ch, " ")
    return cleaned[:31] or "Sheet"
