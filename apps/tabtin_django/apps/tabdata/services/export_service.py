"""
数据导出服务

提供CSV、Excel、JSON和PDF文件的导出功能
支持导出当前视图、选中记录、筛选后的数据
"""
import csv
import logging
import re
import base64
import binascii
from xml.sax.saxutils import escape as xml_escape
import functools
import io
import json
import copy
from datetime import datetime, date, timezone as dt_timezone
from typing import List, Dict, Any, Optional, Iterable, Set, Tuple
from uuid import UUID
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
from django.contrib.auth import get_user_model
from django.db.models import Case, IntegerField, Value, When
from django.utils import timezone
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.units import pixels_to_EMU
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image as PDFImage,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Table as PDFTable,
    TableStyle,
)
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase import pdfdoc
from reportlab.pdfgen.canvas import Canvas

from apps.tabdata.constants import (
    TABDATA_DB_ALIAS, FILE_BASED_FIELD_TYPES,
    MAX_EXPORT_ROWS, MAX_EXPORT_ROWS_PDF,
)
from apps.tabdata.models import AttachmentReference, Table, TableField, TableRecord, TableView
from apps.tabdata.utils.record_data_access import read_data, read_data_bulk
from apps.tabdata.utils.view_serializers import (
    build_view_column_meta,
    build_view_column_meta_payload,
)
from .base import BaseService
from apps.services.oss.services.factory import get_oss_service
from apps.services.oss.models import FileRecord
from apps.services.oss.services.public_assets import public_asset_object_key_from_ref

logger = logging.getLogger(__name__)

User = get_user_model()

PDF_CELL_MAX_LEN = 200

_PDF_HORIZONTAL_MARGIN = 0.45 * inch
_PDF_VERTICAL_MARGIN = 0.5 * inch
_PDF_CELL_HORIZONTAL_PADDING = 8
_PDF_MIN_COLUMN_WIDTH = 0.8 * inch
_PDF_MAX_COLUMN_WIDTH = 2.5 * inch
_PDF_PRIMARY_COLUMN_MIN_WIDTH = 1.35 * inch
_PDF_LONG_TEXT_MIN_WIDTH = 1.35 * inch
_PDF_USER_COLUMN_MIN_WIDTH = 1.1 * inch
_PDF_DATE_COLUMN_MIN_WIDTH = 1.15 * inch
_PDF_LINK_COLUMN_MIN_WIDTH = 1.2 * inch
_PDF_WIDTH_SAMPLE_SIZE = 200
_PDF_ATTACHMENT_MAX_BYTES = 50 * 1024 * 1024
_PDF_ATTACHMENTS_TOTAL_MAX_BYTES = 100 * 1024 * 1024
_PDF_ATTACHMENT_NAME_MAX_LEN = 200

_USER_FIELD_TYPES = frozenset(('user', 'created_by', 'last_modified_by'))
_PDF_DATE_FIELD_TYPES = frozenset(('date', 'created_time', 'last_modified_time'))

_BOOL_LABEL_YES = 'Yes'
_BOOL_LABEL_NO = 'No'

_VIEW_EXPORT_PAGE_SIZE = 1000


_CJK_FONT_CANDIDATES = ('STSong-Light', 'STHeiti-Light', 'MSung-Light', 'HeiseiMin-W3')

_SYSTEM_CJK_FONT_PATHS = (
    '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
    '/usr/share/fonts/opentype/noto/NotoSansCJKsc-Regular.otf',
    '/usr/share/fonts/noto-cjk/NotoSansCJK-Regular.ttc',
    '/usr/share/fonts/google-noto-cjk/NotoSansCJK-Regular.ttc',
    '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc',
    '/usr/share/fonts/OTF/NotoSansCJK-Regular.ttc',
    '/System/Library/Fonts/STHeiti Light.ttc',
    '/System/Library/Fonts/PingFang.ttc',
)


@functools.lru_cache(maxsize=1)
def _get_cjk_font_name() -> str:
    """线程安全的 CJK 字体初始化（lru_cache 内部加锁，保证只执行一次）。"""
    import os

    for font_name in _CJK_FONT_CANDIDATES:
        try:
            pdfmetrics.registerFont(UnicodeCIDFont(font_name))
            return font_name
        except Exception:
            logger.debug('CJK CID 字体 %s 注册失败', font_name, exc_info=True)
            continue

    env_path = os.environ.get('MUSE_CJK_FONT_PATH')
    ttf_search_paths = [env_path] if env_path else list(_SYSTEM_CJK_FONT_PATHS)

    for font_path in ttf_search_paths:
        if font_path and os.path.isfile(font_path):
            try:
                registered_name = 'TabTinCJK'
                pdfmetrics.registerFont(TTFont(registered_name, font_path))
                logger.info('PDF 导出已注册系统 CJK 字体: %s', font_path)
                return registered_name
            except Exception:
                logger.debug('CJK TTF 字体 %s 注册失败', font_path, exc_info=True)
                continue

    logger.warning(
        'PDF 导出无可用 CJK 字体（CID: %s, 系统路径均未找到），中文将显示为方框',
        ', '.join(_CJK_FONT_CANDIDATES),
    )
    return 'Helvetica'


_CSV_INJECTION_CHARS = frozenset('=+-@\t\r\n')


class _Echo:
    """csv.writer 需要的最小 file-like 对象，直接返回写入值用于流式输出。"""
    def write(self, value):
        return value


def _sanitize_cell_value(value):
    """防止 Formula Injection：Excel/CSV 等会将以特殊字符开头的值视为公式执行（OWASP CWE-1236）"""
    if isinstance(value, str) and value and value[0] in _CSV_INJECTION_CHARS:
        return "'" + value
    return value


_sanitize_csv_value = _sanitize_cell_value


_EXCEL_CELL_MAX_LEN = 32000
_EXCEL_IMAGE_MAX_BYTES = 5 * 1024 * 1024
_EXCEL_IMAGE_MAX_WIDTH = 240
_EXCEL_IMAGE_MAX_HEIGHT = 180
_EXCEL_IMAGE_ROW_HEIGHT = 140
_EXCEL_IMAGE_COL_WIDTH = 35
_EXCEL_IMAGE_STACK_GAP_PX = 8
_EXCEL_IMAGE_TEXT_OFFSET_PX = 20
_EXCEL_POINTS_PER_PIXEL = 0.75
_PDF_IMAGE_MAX_WIDTH = 1.4 * inch
_PDF_IMAGE_MAX_HEIGHT = 1.05 * inch
_EXCEL_IMAGE_MIME_PREFIX = 'image/'
_EXCEL_IMAGE_EXTENSIONS = ('.png', '.jpg', '.jpeg', '.gif', '.bmp')
_DATA_IMAGE_URL_RE = re.compile(
    r'^data:(?P<mime>image/[a-z0-9.+-]+);base64,(?P<data>.+)$',
    re.IGNORECASE | re.DOTALL,
)
_DATA_FILE_URL_RE = re.compile(
    r'^data:(?P<mime>[a-z0-9.+-]+/[a-z0-9.+-]+);base64,(?P<data>.+)$',
    re.IGNORECASE | re.DOTALL,
)
_DATE_ONLY_VALUE_RE = re.compile(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$')


def _format_file_field_value(value: Any, hide_image_names: bool = False) -> str:
    """将 attachment 字段格式化为人类可读文本（文件名列表）。

    结果会截断到 _EXCEL_CELL_MAX_LEN 以内，确保 Excel 单元格安全。
    """
    if isinstance(value, dict):
        if hide_image_names and _is_image_file_item(value):
            return ''
        name = value.get('name') or value.get('file_name') or value.get('url') or ''
        return str(name) if name else ''
    if not isinstance(value, list):
        return str(value) if value else ''
    names = []
    for item in value:
        if isinstance(item, dict):
            if hide_image_names and _is_image_file_item(item):
                continue
            name = item.get('name') or item.get('file_name') or item.get('url') or ''
            if name:
                names.append(str(name))
        elif item is not None:
            names.append(str(item))
    result = ', '.join(names) if names else ''
    if len(result) > _EXCEL_CELL_MAX_LEN:
        result = result[:_EXCEL_CELL_MAX_LEN - 3] + '...'
    return result


def _iter_file_field_items(value: Any) -> Iterable[Dict[str, Any]]:
    if isinstance(value, dict):
        yield value
    elif isinstance(value, list):
        for item in value:
            if isinstance(item, dict):
                yield item


def _is_image_file_item(item: Dict[str, Any]) -> bool:
    mime = str(item.get('mime_type') or '').lower().strip()
    if mime.startswith(_EXCEL_IMAGE_MIME_PREFIX):
        return True

    file_type = str(item.get('file_type') or '').lower().strip()
    if file_type == 'image':
        return True

    name = str(item.get('name') or item.get('file_name') or '').lower().strip()
    return name.endswith(_EXCEL_IMAGE_EXTENSIONS)


def _decode_data_image_url(value: str) -> Optional[bytes]:
    match = _DATA_IMAGE_URL_RE.match(value.strip())
    if not match:
        return None

    try:
        image_bytes = base64.b64decode(match.group('data'), validate=True)
    except (binascii.Error, ValueError):
        return None

    if len(image_bytes) > _EXCEL_IMAGE_MAX_BYTES:
        return None
    return image_bytes


def _decode_data_file_url(value: str, max_bytes: int) -> Optional[bytes]:
    match = _DATA_FILE_URL_RE.match(value.strip())
    if not match:
        return None

    # A base64 payload is at least 4/3 of its decoded bytes. Reject oversized
    # values before decoding so a forged/missing cell size cannot allocate an
    # unbounded temporary byte string in the export worker.
    if len(match.group('data')) > ((max_bytes + 2) // 3) * 4:
        return None

    try:
        file_bytes = base64.b64decode(match.group('data'), validate=True)
    except (binascii.Error, ValueError):
        return None
    return file_bytes if len(file_bytes) <= max_bytes else None


def _download_oss_file_bytes(
    object_key: str,
    cache: Dict[str, Optional[bytes]],
) -> Optional[bytes]:
    if not object_key:
        return None
    if object_key in cache:
        return cache[object_key]

    file_bytes: Optional[bytes] = None
    try:
        result = get_oss_service().download_file(object_key)
        data = result.get('data') if result.get('success') else None
        content = data.get('content') if isinstance(data, dict) else None
        if isinstance(content, bytes) and len(content) <= _PDF_ATTACHMENT_MAX_BYTES:
            file_bytes = content
    except Exception:
        logger.debug('PDF export attachment download skipped: object_key=%s', object_key, exc_info=True)

    cache[object_key] = file_bytes
    return file_bytes


def _resolve_authorized_pdf_file(
    item: Dict[str, Any],
    authorized_files: Dict[str, FileRecord],
) -> Optional[FileRecord]:
    file_id = item.get('file_id')
    if file_id:
        file_record = authorized_files.get(f'id:{file_id}')
        if file_record is not None:
            return file_record

    object_key = _resolve_oss_object_key_from_file_item(item)
    if object_key:
        return authorized_files.get(f'key:{object_key}')
    return None


def _resolve_file_bytes_from_file_item(
    item: Dict[str, Any],
    oss_file_cache: Dict[str, Optional[bytes]],
    authorized_files: Dict[str, FileRecord],
) -> Optional[bytes]:
    try:
        declared_size = int(item.get('size') or 0)
    except (TypeError, ValueError):
        declared_size = 0
    if declared_size > _PDF_ATTACHMENT_MAX_BYTES:
        return None

    for url_key in ('url', 'data_url'):
        url_value = item.get(url_key)
        if isinstance(url_value, str):
            file_bytes = _decode_data_file_url(url_value, _PDF_ATTACHMENT_MAX_BYTES)
            if file_bytes is not None:
                return file_bytes

    file_record = _resolve_authorized_pdf_file(item, authorized_files)
    if file_record is not None and file_record.file_size <= _PDF_ATTACHMENT_MAX_BYTES:
        return _download_oss_file_bytes(file_record.file_key, oss_file_cache)
    return None


def _sanitize_pdf_attachment_name(item: Dict[str, Any], fallback_index: int) -> str:
    raw_name = item.get('name') or item.get('file_name')
    if not raw_name:
        object_key = _resolve_oss_object_key_from_file_item(item)
        raw_name = re.split(r'[\\/]+', object_key)[-1] if object_key else ''
    name = re.split(r'[\\/]+', str(raw_name))[-1]
    name = ''.join(character for character in name if ord(character) >= 32).strip(' .')
    if not name:
        name = f'attachment-{fallback_index}'
    return name[:_PDF_ATTACHMENT_NAME_MAX_LEN]


def _dedupe_pdf_attachment_name(name: str, used_names: set[str]) -> str:
    if name not in used_names:
        used_names.add(name)
        return name

    stem, separator, suffix = name.rpartition('.')
    if not separator or not stem:
        stem, suffix = name, ''
    else:
        suffix = f'.{suffix}'
    duplicate_index = 2
    while True:
        marker = f' ({duplicate_index})'
        max_stem_length = _PDF_ATTACHMENT_NAME_MAX_LEN - len(marker) - len(suffix)
        candidate = f'{stem[:max(1, max_stem_length)]}{marker}{suffix}'
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        duplicate_index += 1


def _collect_pdf_attachments(
    fields: List[TableField],
    raw_rows: List[List[Any]],
    raw_record_ids: List[str],
    authorized_files_by_cell: Dict[tuple[str, str], Dict[str, FileRecord]],
    oss_file_cache: Dict[str, Optional[bytes]],
) -> List[tuple[str, bytes]]:
    attachments: List[tuple[str, bytes]] = []
    used_names: set[str] = set()
    total_bytes = 0
    fallback_index = 1
    embedded_file_ids: set[str] = set()

    for row_index, raw_row in enumerate(raw_rows):
        for column_index, field in enumerate(fields):
            if field.field_type not in FILE_BASED_FIELD_TYPES:
                continue
            authorized_files = authorized_files_by_cell.get(
                (raw_record_ids[row_index], str(field.id)),
                {},
            )
            for item in _iter_file_field_items(raw_row[column_index]):
                file_record = _resolve_authorized_pdf_file(item, authorized_files)
                file_identity = str(file_record.id) if file_record is not None else ''
                if file_identity and file_identity in embedded_file_ids:
                    continue
                file_bytes = _resolve_file_bytes_from_file_item(
                    item,
                    oss_file_cache,
                    authorized_files,
                )
                if file_bytes is None:
                    continue
                if total_bytes + len(file_bytes) > _PDF_ATTACHMENTS_TOTAL_MAX_BYTES:
                    logger.warning(
                        'PDF export attachment skipped because total embedded size limit was reached: name=%s',
                        item.get('name') or item.get('file_name') or '',
                    )
                    continue
                name = _sanitize_pdf_attachment_name(item, fallback_index)
                fallback_index += 1
                attachments.append((_dedupe_pdf_attachment_name(name, used_names), file_bytes))
                total_bytes += len(file_bytes)
                if file_identity:
                    embedded_file_ids.add(file_identity)
    return attachments


def _build_pdf_authorized_files_by_cell(
    table: Table,
    fields: List[TableField],
    record_ids: List[str],
) -> Dict[tuple[str, str], Dict[str, FileRecord]]:
    file_field_ids = [field.id for field in fields if field.field_type in FILE_BASED_FIELD_TYPES]
    if not record_ids or not file_field_ids:
        return {}

    references = list(
        AttachmentReference.objects.using(TABDATA_DB_ALIAS).filter(
            table_id=table.id,
            record_id__in=record_ids,
            field_id__in=file_field_ids,
            organization_id=table.organization_id,
            is_deleted=False,
        ).values('record_id', 'field_id', 'file_id')
    )
    if not references:
        return {}

    file_ids = {reference['file_id'] for reference in references}
    file_records = {
        file_record.id: file_record
        for file_record in FileRecord.objects.using('default').filter(
            id__in=file_ids,
            status='completed',
            file_size__lte=_PDF_ATTACHMENT_MAX_BYTES,
        )
    }

    # Match AttachmentService's legacy rule: an organization-less FileRecord
    # is usable only when every active, internally consistent reference belongs
    # to this table's organization.
    legacy_file_ids = {
        file_id
        for file_id, file_record in file_records.items()
        if not file_record.organization_id and not file_record.is_public
    }
    legacy_reference_organizations: Dict[UUID, set[str]] = {
        file_id: set() for file_id in legacy_file_ids
    }
    if legacy_file_ids:
        for reference in AttachmentReference.objects.using(TABDATA_DB_ALIAS).filter(
            file_id__in=legacy_file_ids,
            is_deleted=False,
        ).values('file_id', 'organization_id', 'table__organization_id'):
            if str(reference['organization_id']) == str(reference['table__organization_id']):
                legacy_reference_organizations[reference['file_id']].add(
                    str(reference['table__organization_id'])
                )

    table_organization_id = str(table.organization_id)
    authorized_records: Dict[UUID, FileRecord] = {}
    for file_id, file_record in file_records.items():
        if str(file_record.organization_id or '') == table_organization_id:
            authorized_records[file_id] = file_record
        elif (
            not file_record.organization_id
            and not file_record.is_public
            and legacy_reference_organizations.get(file_id) == {table_organization_id}
        ):
            authorized_records[file_id] = file_record

    authorized_files_by_cell: Dict[tuple[str, str], Dict[str, FileRecord]] = {}
    for reference in references:
        file_record = authorized_records.get(reference['file_id'])
        if file_record is None:
            continue
        cell_files = authorized_files_by_cell.setdefault(
            (str(reference['record_id']), str(reference['field_id'])),
            {},
        )
        cell_files[f'id:{file_record.id}'] = file_record
        cell_files[f'key:{file_record.file_key}'] = file_record
    return authorized_files_by_cell


class _PDFAttachmentCanvas(Canvas):
    def __init__(self, *args, attachments: List[tuple[str, bytes]], **kwargs):
        super().__init__(*args, **kwargs)
        self._tabtin_attachments = attachments

    def save(self):
        if self._tabtin_attachments:
            names = []
            for name, file_bytes in sorted(self._tabtin_attachments, key=lambda attachment: attachment[0]):
                embedded_file = pdfdoc.PDFStream(
                    pdfdoc.PDFDictionary({
                        'Type': pdfdoc.PDFName('EmbeddedFile'),
                        'Params': pdfdoc.PDFDictionary({'Size': len(file_bytes)}),
                    }),
                    file_bytes,
                )
                embedded_file_reference = self._doc.Reference(embedded_file)
                file_specification = pdfdoc.PDFDictionary({
                    'Type': pdfdoc.PDFName('Filespec'),
                    'F': pdfdoc.PDFString(name),
                    'UF': pdfdoc.PDFString(name),
                    'EF': pdfdoc.PDFDictionary({
                        'F': embedded_file_reference,
                        'UF': embedded_file_reference,
                    }),
                })
                names.extend([
                    pdfdoc.PDFString(name),
                    self._doc.Reference(file_specification),
                ])
            self._doc.Catalog.Names = pdfdoc.PDFDictionary({
                'EmbeddedFiles': pdfdoc.PDFDictionary({
                    'Names': pdfdoc.PDFArray(names),
                }),
            })
        super().save()


def _download_oss_image_bytes(object_key: str, cache: Dict[str, Optional[bytes]]) -> Optional[bytes]:
    if not object_key:
        return None
    if object_key in cache:
        return cache[object_key]

    image_bytes: Optional[bytes] = None
    try:
        result = get_oss_service().download_file(object_key)
        data = result.get('data') if result.get('success') else None
        content = data.get('content') if isinstance(data, dict) else None
        content_type = str(data.get('content_type') or '').lower() if isinstance(data, dict) else ''
        if isinstance(content, bytes) and content_type.startswith(_EXCEL_IMAGE_MIME_PREFIX):
            if len(content) <= _EXCEL_IMAGE_MAX_BYTES:
                image_bytes = content
    except Exception:
        logger.debug('Excel export image download skipped: object_key=%s', object_key, exc_info=True)

    cache[object_key] = image_bytes
    return image_bytes


def _resolve_oss_object_key_from_file_item(item: Dict[str, Any]) -> str:
    for key_name in ('key', 'file_key', 'object_key', 'oss_key'):
        value = item.get(key_name)
        if isinstance(value, str) and value.strip():
            return value.strip()

    for url_key in ('url', 'thumbnail_url', 'data_url', 'access_url', 'cdn_url'):
        value = item.get(url_key)
        if not isinstance(value, str):
            continue
        object_key = public_asset_object_key_from_ref(value)
        if object_key:
            return object_key

    return ''


def _resolve_image_bytes_from_file_item(
    item: Dict[str, Any],
    oss_image_cache: Dict[str, Optional[bytes]],
) -> Optional[bytes]:
    if not _is_image_file_item(item):
        return None

    try:
        declared_size = int(item.get('size') or 0)
    except (TypeError, ValueError):
        declared_size = 0
    if declared_size > _EXCEL_IMAGE_MAX_BYTES:
        return None

    for url_key in ('url', 'thumbnail_url', 'data_url'):
        url_value = item.get(url_key)
        if isinstance(url_value, str):
            image_bytes = _decode_data_image_url(url_value)
            if image_bytes:
                return image_bytes

    object_key = _resolve_oss_object_key_from_file_item(item)
    if object_key:
        return _download_oss_image_bytes(object_key, oss_image_cache)

    return None


def _build_excel_image(image_bytes: bytes) -> Optional[OpenpyxlImage]:
    try:
        from PIL import Image as PILImage

        source = io.BytesIO(image_bytes)
        with PILImage.open(source) as pil_image:
            original_width, original_height = pil_image.size
        source.seek(0)

        excel_image = OpenpyxlImage(source)
        if original_width > 0 and original_height > 0:
            display_scale = min(
                _EXCEL_IMAGE_MAX_WIDTH / original_width,
                _EXCEL_IMAGE_MAX_HEIGHT / original_height,
                1,
            )
            excel_image.width = int(original_width * display_scale)
            excel_image.height = int(original_height * display_scale)
        # openpyxl reads from the stream during save, so keep it alive on the image object.
        excel_image._tabtin_stream = source
        return excel_image
    except Exception:
        logger.debug('Excel export image preparation skipped', exc_info=True)
        return None


def _build_excel_image_from_file_field(
    value: Any,
    oss_image_cache: Dict[str, Optional[bytes]],
) -> Optional[OpenpyxlImage]:
    images = _build_excel_images_from_file_field(value, oss_image_cache)
    return images[0] if images else None


def _build_excel_images_from_file_field(
    value: Any,
    oss_image_cache: Dict[str, Optional[bytes]],
) -> List[OpenpyxlImage]:
    images: List[OpenpyxlImage] = []
    for item in _iter_file_field_items(value):
        image_bytes = _resolve_image_bytes_from_file_item(item, oss_image_cache)
        if image_bytes:
            image = _build_excel_image(image_bytes)
            if image:
                images.append(image)
    return images


def _build_excel_image_anchor(
    row_number: int,
    col_number: int,
    image: OpenpyxlImage,
    y_offset_px: int = 0,
) -> OneCellAnchor:
    marker = AnchorMarker(
        col=col_number - 1,
        colOff=pixels_to_EMU(2),
        row=row_number - 1,
        rowOff=pixels_to_EMU(y_offset_px),
    )
    return OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(
            cx=pixels_to_EMU(image.width),
            cy=pixels_to_EMU(image.height),
        ),
    )


def _build_pdf_image(image_bytes: bytes, max_width: float, max_height: float) -> Optional[PDFImage]:
    try:
        from PIL import Image as PILImage

        source = io.BytesIO(image_bytes)
        with PILImage.open(source) as pil_image:
            original_width, original_height = pil_image.size
        source.seek(0)

        display_width = float(original_width)
        display_height = float(original_height)
        if original_width > 0 and original_height > 0:
            display_scale = min(
                max_width / original_width,
                max_height / original_height,
                1,
            )
            display_width = max(1.0, original_width * display_scale)
            display_height = max(1.0, original_height * display_scale)

        pdf_image = PDFImage(source, width=display_width, height=display_height)
        # ReportLab reads from the stream during doc.build(), so keep it alive.
        pdf_image._tabtin_stream = source
        return pdf_image
    except Exception:
        logger.debug('PDF export image preparation skipped', exc_info=True)
        return None


def _build_pdf_image_from_file_field(
    value: Any,
    oss_image_cache: Dict[str, Optional[bytes]],
    authorized_files: Dict[str, FileRecord],
    max_width: float,
    max_height: float = _PDF_IMAGE_MAX_HEIGHT,
) -> Optional[PDFImage]:
    for item in _iter_file_field_items(value):
        if not _is_image_file_item(item):
            continue
        image_bytes = None
        for url_key in ('url', 'thumbnail_url', 'data_url'):
            url_value = item.get(url_key)
            if isinstance(url_value, str):
                image_bytes = _decode_data_image_url(url_value)
                if image_bytes:
                    break
        if image_bytes is None:
            file_record = _resolve_authorized_pdf_file(item, authorized_files)
            if file_record is not None and file_record.file_size <= _EXCEL_IMAGE_MAX_BYTES:
                image_bytes = _download_oss_image_bytes(file_record.file_key, oss_image_cache)
        if image_bytes:
            image = _build_pdf_image(image_bytes, max_width, max_height)
            if image:
                return image
    return None


def _format_cell_value(
    value: Any,
    field_type: str = '',
    bool_style: str = 'bool',
    hide_file_image_names: bool = False,
) -> str:
    """统一的单元格值格式化。

    bool_style:
      - 'bool': true/false（CSV 风格）
      - 'label': Yes/No（Excel/PDF 风格）
    """
    if field_type in FILE_BASED_FIELD_TYPES and isinstance(value, (list, dict)):
        return _format_file_field_value(value, hide_image_names=hide_file_image_names)
    if field_type == 'multi_select' and isinstance(value, list):
        return ', '.join(str(v) for v in value)
    if field_type in ('user', 'created_by', 'last_modified_by'):
        if isinstance(value, dict):
            return value.get('name', str(value))
        if isinstance(value, list):
            names = []
            for item in value:
                if isinstance(item, dict):
                    names.append(item.get('name', str(item)))
                elif item is not None:
                    names.append(str(item))
            return ', '.join(names) if names else ''
    if field_type == 'link' and isinstance(value, dict):
        link_title = value.get('title') or value.get('name') or value.get('id', '')
        return str(link_title) if link_title else ''
    if field_type == 'link' and isinstance(value, list):
        titles = []
        for item in value:
            if isinstance(item, dict):
                t = item.get('title') or item.get('name') or item.get('id', '')
                if t:
                    titles.append(str(t))
            elif item is not None:
                titles.append(str(item))
        return ', '.join(titles) if titles else ''
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        if bool_style == 'label':
            return _BOOL_LABEL_YES if value else _BOOL_LABEL_NO
        return 'true' if value else 'false'
    return str(value)


def _apply_display_timezone(value: datetime, time_zone: str) -> datetime:
    if not time_zone:
        return value
    try:
        target_zone = ZoneInfo(time_zone)
    except (ZoneInfoNotFoundError, ValueError):
        return value
    aware_value = value if value.tzinfo else value.replace(tzinfo=dt_timezone.utc)
    return aware_value.astimezone(target_zone)


def _resolve_date_parts(value: Any, time_zone: str = '') -> Optional[Dict[str, Any]]:
    if not value:
        return None
    if isinstance(value, datetime):
        value = _apply_display_timezone(value, time_zone)
        return {
            'year': value.year,
            'month': value.month,
            'day': value.day,
            'hour': value.hour,
            'minute': value.minute,
        }
    if isinstance(value, date):
        return {
            'year': value.year,
            'month': value.month,
            'day': value.day,
            'hour': None,
            'minute': None,
        }
    if isinstance(value, str):
        stripped = value.strip()
        match = _DATE_ONLY_VALUE_RE.match(stripped)
        if match:
            return {
                'year': int(match.group(1)),
                'month': int(match.group(2)),
                'day': int(match.group(3)),
                'hour': None,
                'minute': None,
            }
        try:
            parsed = datetime.fromisoformat(stripped.replace('Z', '+00:00'))
            parsed = _apply_display_timezone(parsed, time_zone)
            return {
                'year': parsed.year,
                'month': parsed.month,
                'day': parsed.day,
                'hour': parsed.hour,
                'minute': parsed.minute,
            }
        except ValueError:
            return None
    return None


def _format_date_parts(parts: Dict[str, int], date_format: str) -> str:
    year = str(parts['year'])
    month = f"{parts['month']:02d}"
    day = f"{parts['day']:02d}"
    month_numeric = str(parts['month'])
    day_numeric = str(parts['day'])
    if date_format == 'YYYY-MM-DD':
        return f'{year}-{month}-{day}'
    if date_format == 'M/D/YYYY':
        return f'{month_numeric}/{day_numeric}/{year}'
    if date_format == 'D/M/YYYY':
        return f'{day_numeric}/{month_numeric}/{year}'
    if date_format == 'YYYY/MM/DD':
        return f'{year}/{month}/{day}'
    # Match the TabData grid default zh-CN-style date display.
    return f'{year}/{month_numeric}/{day_numeric}'


def _format_time_parts(parts: Dict[str, Any], time_format: str) -> str:
    hour = parts['hour']
    minute = f"{parts['minute']:02d}"
    if time_format == 'hh:mm A':
        meridiem = 'PM' if hour >= 12 else 'AM'
        hour12 = hour % 12 or 12
        return f'{hour12:02d}:{minute} {meridiem}'
    return f'{hour:02d}:{minute}'


def _format_excel_date_cell_value(value: Any, field: TableField) -> str:
    options = field.config if isinstance(field.config, dict) else {}
    formatting = options.get('formatting') if isinstance(options.get('formatting'), dict) else {}
    date_format = formatting.get('date') if isinstance(formatting.get('date'), str) else ''
    time_format = formatting.get('time') if isinstance(formatting.get('time'), str) else ''
    time_zone = formatting.get('timeZone') if isinstance(formatting.get('timeZone'), str) else ''

    parts = _resolve_date_parts(value, time_zone)
    if not parts:
        return _format_cell_value(value, field.field_type, bool_style='label')

    date_text = _format_date_parts(parts, date_format)
    if time_format == 'None' or parts['hour'] is None or parts['minute'] is None:
        return date_text
    if time_format in ('HH:mm', 'hh:mm A'):
        return f"{date_text} {_format_time_parts(parts, time_format)}"
    return date_text


_SYSTEM_USER_FIELD_TYPES = frozenset(('created_by', 'last_modified_by'))


def _extract_user_value_parts(value: Any) -> tuple[str, str]:
    """返回用户字段项的 ``(user_id, embedded_name)``，兼容历史存储形态。"""
    if isinstance(value, dict):
        raw_id = value.get('id') or value.get('user_id') or value.get('userId')
        raw_name = (
            value.get('name')
            or value.get('display_name')
            or value.get('displayName')
            or value.get('email')
        )
        return str(raw_id or '').strip(), str(raw_name or '').strip()
    if value is None:
        return '', ''
    return str(value).strip(), ''


def _collect_user_ids(value: Any) -> set[str]:
    values = value if isinstance(value, list) else [value]
    return {
        user_id
        for item in values
        for user_id, _embedded_name in [_extract_user_value_parts(item)]
        if user_id
    }


def _build_pdf_user_display_name_map(
    organization_id: UUID,
    user_ids: set[str],
) -> Tuple[Dict[str, str], Set[str]]:
    """
    按 Organization 边界批量解析人员字段的展示名。

    返回 ``(display_name_by_id, departed_user_ids)``。姓名表把在职成员和离组快照合在一起
    方便查询，但导出时要区分两者——离组的人得标出来，否则读者会以为他还在岗。
    """
    if not organization_id or not user_ids:
        return {}, set()

    valid_user_ids: set[UUID] = set()
    for user_id in user_ids:
        try:
            valid_user_ids.add(UUID(user_id))
        except (TypeError, ValueError, AttributeError):
            continue
    if not valid_user_ids:
        return {}, set()

    from apps.tabtinspace.models import (
        OrganizationMember,
        OrganizationMemberIdentitySnapshot,
    )

    current_member_ids = set(
        OrganizationMember.objects.filter(
            organization_id=organization_id,
            user_id__in=valid_user_ids,
        ).values_list('user_id', flat=True)
    )
    display_names = {
        str(user.id): user.get_display_name()
        for user in User.objects.filter(id__in=current_member_ids).only(
            'id', 'nickname', 'username', 'email', 'phone'
        )
    }

    departed_user_ids: Set[str] = set()
    unresolved_ids = valid_user_ids - current_member_ids
    if unresolved_ids:
        historical_names = OrganizationMemberIdentitySnapshot.objects.filter(
            organization_id=organization_id,
            user_id__in=unresolved_ids,
        ).values_list('user_id', 'display_name')
        for user_id, display_name in historical_names:
            normalized_name = str(display_name or '').strip()
            if normalized_name:
                display_names[str(user_id)] = normalized_name
                departed_user_ids.add(str(user_id))

    return display_names, departed_user_ids


# 人员字段导出文案，与桌面 Grid、移动端保持同一口径。
_DEPARTED_MEMBER_SUFFIX = '（已离职）'
_UNKNOWN_MEMBER_LABEL = '未知'


def _format_pdf_user_value(
    value: Any,
    display_name_by_id: Dict[str, str],
    departed_user_ids: Set[str] | None = None,
) -> str:
    """
    把 user 类字段转换为 UI 同口径的可读姓名。

    查不到的 ID 说「未知」而不是「已离开成员(前8位)」：这个 ID 可能从来没对应过任何用户
    （脏数据、误写、跨组织残留），断言他"离开过"是在编造事实；露出 ID 片段又把内部标识
    带进了可分发的导出文件。
    """
    departed_ids = departed_user_ids or set()
    values = value if isinstance(value, list) else [value]
    display_names: List[str] = []
    for item in values:
        user_id, embedded_name = _extract_user_value_parts(item)
        if not user_id and not embedded_name:
            continue
        resolved_name = display_name_by_id.get(user_id)
        if resolved_name:
            display_names.append(
                f'{resolved_name}{_DEPARTED_MEMBER_SUFFIX}'
                if user_id in departed_ids
                else resolved_name
            )
        elif embedded_name:
            display_names.append(embedded_name)
        else:
            display_names.append(_UNKNOWN_MEMBER_LABEL)
    return ', '.join(display_names)


def _pdf_column_min_width(field: TableField) -> float:
    if field.is_primary:
        return _PDF_PRIMARY_COLUMN_MIN_WIDTH
    if field.field_type == 'long_text':
        return _PDF_LONG_TEXT_MIN_WIDTH
    if field.field_type in _USER_FIELD_TYPES:
        return _PDF_USER_COLUMN_MIN_WIDTH
    if field.field_type in _PDF_DATE_FIELD_TYPES:
        return _PDF_DATE_COLUMN_MIN_WIDTH
    if field.field_type in ('url', 'email', 'phone', 'link'):
        return _PDF_LINK_COLUMN_MIN_WIDTH
    return _PDF_MIN_COLUMN_WIDTH


def _measure_pdf_column_widths(
    fields: List[TableField],
    formatted_rows: List[List[str]],
    font_name: str,
    view_width_by_field_id: Dict[str, int],
) -> List[float]:
    """结合当前视图宽度和实际字体度量，计算每列的可读目标宽度。"""
    column_widths: List[float] = []
    sampled_rows = formatted_rows[:_PDF_WIDTH_SAMPLE_SIZE]
    for column_index, field in enumerate(fields):
        measured_widths: List[float] = []
        texts = [field.name, *(row[column_index] for row in sampled_rows)]
        for text in texts:
            try:
                measured_widths.append(pdfmetrics.stringWidth(text, font_name, 7))
            except (KeyError, TypeError, ValueError):
                measured_widths.append(len(text) * 7)

        measured_widths.sort()
        percentile_index = max(0, int((len(measured_widths) - 1) * 0.9))
        measured_target = measured_widths[percentile_index] + _PDF_CELL_HORIZONTAL_PADDING
        view_width_px = view_width_by_field_id.get(str(field.id), field.width or 0)
        view_target = max(0, view_width_px) * 0.75
        column_widths.append(
            min(
                _PDF_MAX_COLUMN_WIDTH,
                max(_pdf_column_min_width(field), measured_target, view_target),
            )
        )
    return column_widths


def _partition_pdf_columns(
    fields: List[TableField],
    column_widths: List[float],
    available_width: float,
) -> List[List[int]]:
    """将宽表顺序切成多个横向分片，并在后续分片重复主字段。"""
    if not fields:
        return []
    if sum(column_widths) <= available_width:
        return [list(range(len(fields)))]

    primary_index = next((index for index, field in enumerate(fields) if field.is_primary), None)
    remaining_indexes = [index for index in range(len(fields)) if index != primary_index]
    column_groups: List[List[int]] = []
    cursor = 0
    while cursor < len(remaining_indexes):
        group = [primary_index] if primary_index is not None else []
        group_width = column_widths[primary_index] if primary_index is not None else 0.0
        while cursor < len(remaining_indexes):
            column_index = remaining_indexes[cursor]
            next_width = min(column_widths[column_index], available_width)
            if group and group_width + next_width > available_width:
                break
            group.append(column_index)
            group_width += next_width
            cursor += 1
        if len(group) == (1 if primary_index is not None else 0):
            column_index = remaining_indexes[cursor]
            group.append(column_index)
            cursor += 1
        column_groups.append(sorted(group))

    return column_groups


def _resolve_system_user_value(
    value: Any,
    field_type: str,
    record: Any,
    user_cache: Dict[str, str],
) -> Any:
    """当 created_by / last_modified_by 字段值为 None 时，从 record 元数据查询用户名。"""
    if value is not None or field_type not in _SYSTEM_USER_FIELD_TYPES:
        return value
    uid = None
    if field_type == 'created_by':
        uid = getattr(record, 'created_by_id', None)
    elif field_type == 'last_modified_by':
        uid = getattr(record, 'updated_by_id', None)
    if uid is None:
        return value
    uid_str = str(uid)
    if uid_str not in user_cache:
        try:
            user_obj = User.objects.filter(id=uid).only('id', 'nickname', 'username').first()
            if user_obj:
                user_cache[uid_str] = (
                    getattr(user_obj, 'nickname', None)
                    or getattr(user_obj, 'username', '')
                    or uid_str
                )
            else:
                user_cache[uid_str] = uid_str
        except Exception:
            user_cache[uid_str] = uid_str
    display_name = user_cache.get(uid_str, '')
    return {'id': uid_str, 'name': display_name} if display_name else value


class ExportService(BaseService):
    """
    数据导出服务

    支持导出为CSV、Excel、JSON、PDF格式
    支持视图导出、选中记录导出、筛选导出
    """

    @staticmethod
    def _apply_rls_to_queryset(queryset, table_id, rls_context):
        """通过 native SQL 获取 RLS 允许的记录 ID，过滤 ORM queryset。"""
        from apps.tabdata.services.rls_service import apply_rls_to_orm_queryset
        return apply_rls_to_orm_queryset(queryset, table_id, rls_context)

    @staticmethod
    def _get_export_field_value(record_data: Dict[str, Any], field: TableField, default: Any = '') -> Any:
        """按当前 native/协作存储口径读取字段值，兼容 UUID 与 hex 两类 key。"""
        uuid_key = str(field.id)
        if uuid_key in record_data:
            return record_data[uuid_key]
        hex_key = field.id.hex if isinstance(field.id, UUID) else uuid_key.replace('-', '')
        if hex_key in record_data:
            return record_data[hex_key]
        return default

    @staticmethod
    def _iter_records_with_export_data(
        records_queryset,
        table: Table,
        fields: List[TableField],
        chunk_size: int = 500,
    ) -> Iterable[TableRecord]:
        """分批预加载 native 列，让导出与列表 API 使用同一份持久化数据。"""
        batch: List[TableRecord] = []

        def _flush_batch():
            if batch:
                read_data_bulk(batch, table, fields=fields)
                for item in batch:
                    yield item

        for record in records_queryset.iterator(chunk_size=chunk_size):
            batch.append(record)
            if len(batch) >= chunk_size:
                yield from _flush_batch()
                batch.clear()

        if batch:
            yield from _flush_batch()

    @staticmethod
    def _to_export_field_type(field_type: str) -> str:
        mapping = {
            'text': 'singleLineText',
            'long_text': 'longText',
            'select': 'singleSelect',
            'multi_select': 'multipleSelect',
            'created_time': 'createdTime',
            'last_modified_time': 'lastModifiedTime',
            'created_by': 'createdBy',
            'last_modified_by': 'lastModifiedBy',
        }
        return mapping.get(field_type, field_type)

    def _build_table_full_payload(
        self,
        table: Table,
        records_queryset=None,
    ) -> Dict[str, Any]:
        """
        构造 table_full 结构化快照（dict）。
        """
        all_fields = list(
            TableField.objects.using(TABDATA_DB_ALIAS).filter(
                table_id=table.id,
                is_deleted=False,
            ).order_by('order')
        )
        all_views = list(
            TableView.objects.using(TABDATA_DB_ALIAS).filter(
                table_id=table.id,
            ).order_by('order', 'created_at')
        )

        if records_queryset is None:
            records_queryset = self._get_records_queryset(table.id)

        fields_payload: List[Dict[str, Any]] = []
        for field in all_fields:
            field_config = field.config or {}
            fields_payload.append({
                'id': str(field.id),
                'table_id': str(field.table_id),
                'name': field.name,
                'field_type': field.field_type,
                'type': self._to_export_field_type(field.field_type),
                'description': field.description,
                'config': field_config,
                'options': copy.deepcopy(field_config),
                'validation_rules': field.validation_rules or {},
                'order': field.order,
                'width': field.width,
                'is_primary': bool(field.is_primary),
                'isPrimary': bool(field.is_primary),
                'is_hidden': bool(field.is_hidden),
                'notNull': False,
                'default_value': field.default_value,
                'created_at': field.created_at.isoformat() if field.created_at else None,
                'updated_at': field.updated_at.isoformat() if field.updated_at else None,
            })

        views_payload: List[Dict[str, Any]] = []
        for view in all_views:
            views_payload.append({
                'id': str(view.id),
                'table_id': str(view.table_id),
                'name': view.name,
                'view_type': view.view_type,
                'type': view.view_type,
                'description': view.description or '',
                'config': view.config or {},
                'options': view.config or {},
                'filters': view.filters or [],
                'filter': view.filters or [],
                'sorts': view.sorts or [],
                'sort': view.sorts or [],
                'groups': view.groups or [],
                'group': view.groups or [],
                'visible_fields': view.visible_fields or [],
                'field_order': view.field_order or [],
                **build_view_column_meta_payload(view, table_fields=all_fields),
                'is_shared': bool(view.is_shared),
                'enableShare': bool(view.is_shared),
                'is_locked': bool(view.is_locked),
                'isLocked': bool(view.is_locked),
                'order': view.order,
                'created_by_id': str(view.created_by_id) if view.created_by_id else None,
                'created_at': view.created_at.isoformat() if view.created_at else None,
                'updated_at': view.updated_at.isoformat() if view.updated_at else None,
            })

        records_payload: List[Dict[str, Any]] = []
        skipped_count = 0
        missing_value = object()
        for record in self._iter_records_with_export_data(records_queryset, table, all_fields):
            try:
                source_data = read_data(record)
                normalized_fields = {
                    str(field.id): value
                    for field in all_fields
                    for value in [self._get_export_field_value(source_data, field, missing_value)]
                    if value is not missing_value
                }
                for key, value in source_data.items():
                    if key not in normalized_fields:
                        normalized_fields[key] = value

                records_payload.append({
                    'id': str(record.id),
                    'row_id': str(record.id),
                    'table_id': str(record.table_id),
                    'fields': normalized_fields,
                    'data': normalized_fields,
                    'order': record.order,
                    'version': record.version,
                    'created_by_id': str(record.created_by_id) if record.created_by_id else None,
                    'updated_by_id': str(record.updated_by_id) if record.updated_by_id else None,
                    'created_at': record.created_at.isoformat() if record.created_at else None,
                    'updated_at': record.updated_at.isoformat() if record.updated_at else None,
                })
            except Exception as e:
                logger.warning('导出跳过记录 %s (table %s): %s', record.id, table.id, e)
                skipped_count += 1

        return {
            'id': str(table.id),
            'space_id': str(table.space_id) if table.space_id else None,
            'name': table.name,
            'description': table.description or '',
            'icon': table.icon or '',
            'owner_id': str(table.owner_id) if table.owner_id else None,
            'default_view_id': str(table.default_view_id) if table.default_view_id else None,
            'defaultViewId': str(table.default_view_id) if table.default_view_id else None,
            'row_count': table.row_count,
            'field_count': table.field_count,
            'created_at': table.created_at.isoformat() if table.created_at else None,
            'updated_at': table.updated_at.isoformat() if table.updated_at else None,
            'fields': fields_payload,
            'views': views_payload,
            'records': records_payload,
            'metadata': {
                'format': 'table_full',
                'format_version': 'tabtin.table_full.v1',
                'record_count': len(records_payload),
                'skipped_records': skipped_count,
                'field_count': len(fields_payload),
                'view_count': len(views_payload),
                'export_time': timezone.now().isoformat(),
            },
        }

    def export_space_to_json(
        self,
        space_id: UUID,
        table_ids: Optional[List[UUID]] = None,
        include_archived: bool = False,
        format_type: str = 'base_full',
        rls_context=None,
    ) -> str:
        """
        导出 Space 级快照（BaseJson v1）。

        Args:
            space_id: Space ID
            table_ids: 指定导出的表ID列表（为空则导出全部表）
            include_archived: 是否包含归档表
            format_type: 仅支持 base_full

        Returns:
            str: JSON内容

        Raises:
            PermissionError: 无查看权限或 Space 不存在时抛出
            Space.DoesNotExist: 竞态场景下 Space 被删除时可能抛出
        """
        if format_type != 'base_full':
            raise ValueError(f"不支持的JSON格式类型: {format_type}")

        if not self.check_space_permission(str(space_id), 'viewer'):
            raise PermissionError("没有 Space 的查看权限")

        from apps.tabtinspace.models import Space
        from apps.tabtinspace.services.host_resolver import resolve_host

        space = resolve_host(space_id)
        if space is None:
            raise Space.DoesNotExist

        tables_query = Table.objects.using(TABDATA_DB_ALIAS).filter(space_id=space_id)
        if not include_archived:
            tables_query = tables_query.filter(is_archived=False)
        if table_ids:
            tables_query = tables_query.filter(id__in=table_ids)

        tables = list(tables_query.order_by('created_at', 'id'))
        tables_payload: List[Dict[str, Any]] = []
        for table in tables:
            records_qs = self._get_records_queryset(table.id, max_rows=MAX_EXPORT_ROWS, rls_context=rls_context)
            table_payload = self._build_table_full_payload(table=table, records_queryset=records_qs)
            tables_payload.append(table_payload)

        result = {
            'id': str(space.id),
            'organization_id': str(space.organization_id),
            'name': space.name,
            'description': space.description or '',
            'icon': space.icon or '',
            'color': space.color or '',
            'status': space.status,
            'is_archived': bool(space.is_archived),
            'created_at': space.created_at.isoformat() if space.created_at else None,
            'updated_at': space.updated_at.isoformat() if space.updated_at else None,
            'tables': tables_payload,
            'metadata': {
                'format': 'base_full',
                'format_version': 'tabtin.base_full.v1',
                'table_count': len(tables_payload),
                'export_time': timezone.now().isoformat(),
            },
        }

        return json.dumps(result, ensure_ascii=False, indent=2)

    def _get_records_queryset(
        self,
        table_id: UUID,
        record_ids: Optional[List[UUID]] = None,
        view_id: Optional[UUID] = None,
        filters: Optional[List[Dict[str, Any]]] = None,
        filter_logic: Optional[str] = None,
        sorts: Optional[List[Dict[str, Any]]] = None,
        groups: Optional[List[Dict[str, Any]]] = None,
        max_rows: Optional[int] = None,
        rls_context=None,
    ):
        """
        获取要导出的记录查询集

        Args:
            table_id: 表格ID
            record_ids: 指定的记录ID列表（优先级最高）
            view_id: 视图ID（应用视图的筛选和排序）
            filters: 自定义筛选条件
            max_rows: 最大返回行数（防 OOM），None 表示不限制

        Returns:
            QuerySet: 记录查询集
        """
        queryset = TableRecord.objects.using(TABDATA_DB_ALIAS).filter(
            table_id=table_id,
            is_deleted=False
        )

        # ── RLS 行级安全策略注入 ──
        if rls_context is not None:
            queryset = self._apply_rls_to_queryset(queryset, table_id, rls_context)

        if record_ids:
            ordered_record_ids = list(dict.fromkeys(record_ids))
            preserved_order = Case(
                *[
                    When(id=record_id, then=Value(index))
                    for index, record_id in enumerate(ordered_record_ids)
                ],
                default=Value(len(ordered_record_ids)),
                output_field=IntegerField(),
            )
            queryset = queryset.filter(id__in=ordered_record_ids).order_by(preserved_order)
            if max_rows is not None:
                queryset = queryset[:max_rows]
            return queryset
        elif view_id:
            try:
                view = TableView.objects.using(TABDATA_DB_ALIAS).get(
                    id=view_id,
                    table_id=table_id,
                )
                ordered_record_ids = self._get_view_record_ids(
                    view,
                    max_rows=max_rows,
                    rls_context=rls_context,
                    filters=filters,
                    filter_logic=filter_logic,
                    sorts=sorts,
                    groups=groups,
                )
                if not ordered_record_ids:
                    return queryset.none()

                preserved_order = Case(
                    *[
                        When(id=record_id, then=Value(index))
                        for index, record_id in enumerate(ordered_record_ids)
                    ],
                    default=Value(len(ordered_record_ids)),
                    output_field=IntegerField(),
                )
                return queryset.filter(id__in=ordered_record_ids).order_by(preserved_order)
            except TableView.DoesNotExist:
                raise ValueError("指定的视图不存在或不属于当前表格")

        queryset = queryset.order_by('created_at')
        if max_rows is not None:
            queryset = queryset[:max_rows]
        return queryset

    def _get_view_record_ids(
        self,
        view: TableView,
        *,
        max_rows: Optional[int],
        rls_context=None,
        filters: Optional[List[Dict[str, Any]]] = None,
        filter_logic: Optional[str] = None,
        sorts: Optional[List[Dict[str, Any]]] = None,
        groups: Optional[List[Dict[str, Any]]] = None,
    ) -> List[str]:
        """沿视图共享的 native 数据核心分页收集导出记录 ID。"""
        from .view_data_service import ViewDataService

        service = ViewDataService(user=self.user)
        record_ids: List[str] = []
        seen_ids = set()
        page = 1

        while max_rows is None or len(record_ids) < max_rows:
            remaining = None if max_rows is None else max_rows - len(record_ids)
            page_size = (
                _VIEW_EXPORT_PAGE_SIZE
                if remaining is None
                else min(_VIEW_EXPORT_PAGE_SIZE, remaining)
            )
            if page_size <= 0:
                break

            payload = service._get_grid_data(
                view,
                page=page,
                page_size=page_size,
                fields=set(),
                field_key_type='id',
                rls_context=rls_context,
                filters=filters,
                filter_logic=filter_logic,
                sorts=sorts,
                groups=groups,
            )
            page_records = payload.get('records') or []
            added_count = 0
            for record in page_records:
                record_id = record.get('id') or record.get('row_id')
                if not record_id:
                    continue
                normalized_id = str(record_id)
                if normalized_id in seen_ids:
                    continue
                seen_ids.add(normalized_id)
                record_ids.append(normalized_id)
                added_count += 1
                if max_rows is not None and len(record_ids) >= max_rows:
                    break

            if len(page_records) < page_size or added_count == 0:
                break
            page += 1

        return record_ids

    def _get_export_fields(
        self,
        table_id: UUID,
        field_ids: Optional[List[UUID]] = None,
        view_id: Optional[UUID] = None
    ) -> List[TableField]:
        """
        获取要导出的字段列表

        Args:
            table_id: 表格ID
            field_ids: 指定的字段ID列表
            view_id: 视图ID（使用视图的可见字段）

        Returns:
            List[TableField]: 字段列表
        """
        fields_query = TableField.objects.using(TABDATA_DB_ALIAS).filter(
            table_id=table_id,
            is_deleted=False  # 过滤已删除的字段
        ).exclude(is_hidden=True)

        # 1. 如果指定了字段ID
        if field_ids:
            ordered_field_ids = list(dict.fromkeys(field_ids))
            preserved_order = Case(
                *[
                    When(id=field_id, then=Value(index))
                    for index, field_id in enumerate(ordered_field_ids)
                ],
                default=Value(len(ordered_field_ids)),
                output_field=IntegerField(),
            )
            return list(
                fields_query
                .filter(id__in=ordered_field_ids)
                .order_by(preserved_order)
            )

        # 2. 如果指定了视图，使用视图的可见字段
        elif view_id:
            try:
                view = TableView.objects.using(TABDATA_DB_ALIAS).get(
                    id=view_id,
                    table_id=table_id,
                )
                all_fields = list(fields_query.order_by('order'))
                if not all_fields:
                    return []

                column_meta = build_view_column_meta(view, table_fields=all_fields)
                if not column_meta:
                    return all_fields

                use_visible = any(isinstance(meta, dict) and ('visible' in meta) for meta in column_meta.values())
                use_hidden = any(isinstance(meta, dict) and ('hidden' in meta) for meta in column_meta.values()) and not use_visible

                field_map = {str(field.id): field for field in all_fields}
                default_order_map = {str(field.id): index for index, field in enumerate(all_fields)}

                def _resolve_order(field_id: str) -> float:
                    meta = column_meta.get(field_id) or {}
                    raw_order = meta.get('order') if isinstance(meta, dict) else None
                    if isinstance(raw_order, bool):
                        return float(default_order_map[field_id])
                    if isinstance(raw_order, (int, float)):
                        return float(raw_order)
                    return float(default_order_map[field_id])

                ordered_field_ids = sorted(
                    field_map.keys(),
                    key=lambda field_id: (
                        _resolve_order(field_id),
                        default_order_map[field_id],
                    ),
                )

                visible_fields: List[TableField] = []
                for field_id in ordered_field_ids:
                    field = field_map.get(field_id)
                    if not field:
                        continue
                    meta = column_meta.get(field_id) or {}
                    if use_visible:
                        if isinstance(meta, dict) and 'visible' in meta:
                            if meta.get('visible') is not True:
                                continue
                        elif isinstance(meta, dict) and bool(meta.get('hidden')):
                            continue
                    if use_hidden and bool(meta.get('hidden')):
                        continue
                    visible_fields.append(field)

                if visible_fields:
                    return visible_fields
                return all_fields
            except TableView.DoesNotExist:
                raise ValueError("指定的视图不存在或不属于当前表格")

        return list(fields_query.order_by('order'))

    @staticmethod
    def _format_csv_cell(value: Any, field_type: str = '') -> Any:
        formatted = _format_cell_value(
            value,
            field_type,
            bool_style='bool',
        )
        return _sanitize_csv_value(formatted)

    def export_to_csv(
        self,
        table_id: UUID,
        field_ids: Optional[List[UUID]] = None,
        record_ids: Optional[List[UUID]] = None,
        view_id: Optional[UUID] = None,
        include_headers: bool = True,
        rls_context=None,
        view_query: Optional[Dict[str, Any]] = None,
    ) -> str:
        """
        导出表格数据为CSV（全量加载到内存，适用于小表或需要返回字符串的场景）

        Returns:
            str: CSV内容

        Raises:
            PermissionError: 无查看权限时抛出
            ValueError: 没有可导出的字段时抛出
        """
        if not self.check_table_permission(str(table_id), 'viewer'):
            raise PermissionError("没有表格的查看权限")

        fields = self._get_export_fields(table_id, field_ids, view_id)
        if not fields:
            raise ValueError("没有可导出的字段")

        table = Table.objects.using(TABDATA_DB_ALIAS).get(id=table_id)
        records = self._get_records_queryset(
            table_id, record_ids, view_id, max_rows=MAX_EXPORT_ROWS,
            rls_context=rls_context, **(view_query or {}),
        )
        has_system_user_fields = any(f.field_type in _SYSTEM_USER_FIELD_TYPES for f in fields)
        user_cache: Dict[str, str] = {}

        csv_buffer = io.StringIO()
        csv_buffer.write('\ufeff')
        writer = csv.writer(csv_buffer)

        if include_headers:
            writer.writerow([_sanitize_csv_value(field.name) for field in fields])

        for record in self._iter_records_with_export_data(records, table, fields):
            record_data = read_data(record)
            row = []
            for field in fields:
                raw_value = self._get_export_field_value(record_data, field, '')
                if has_system_user_fields and field.field_type in _SYSTEM_USER_FIELD_TYPES:
                    raw_value = _resolve_system_user_value(raw_value, field.field_type, record, user_cache)
                row.append(self._format_csv_cell(raw_value, field.field_type))
            writer.writerow(row)

        return csv_buffer.getvalue()

    def export_to_csv_streaming(
        self,
        table_id: UUID,
        field_ids: Optional[List[UUID]] = None,
        record_ids: Optional[List[UUID]] = None,
        view_id: Optional[UUID] = None,
        include_headers: bool = True,
        rls_context=None,
        view_query: Optional[Dict[str, Any]] = None,
    ):
        """
        流式导出 CSV，返回可迭代的生成器，配合 StreamingHttpResponse 使用。

        Returns:
            Generator[str]: CSV 行的生成器

        Raises:
            PermissionError: 无查看权限时抛出
            ValueError: 没有可导出的字段时抛出
        """
        if not self.check_table_permission(str(table_id), 'viewer'):
            raise PermissionError("没有表格的查看权限")

        fields = self._get_export_fields(table_id, field_ids, view_id)
        if not fields:
            raise ValueError("没有可导出的字段")

        table = Table.objects.using(TABDATA_DB_ALIAS).get(id=table_id)
        records = self._get_records_queryset(
            table_id, record_ids, view_id, max_rows=MAX_EXPORT_ROWS,
            rls_context=rls_context, **(view_query or {}),
        )
        has_system_user_fields = any(f.field_type in _SYSTEM_USER_FIELD_TYPES for f in fields)
        user_cache: Dict[str, str] = {}
        pseudo_buffer = _Echo()
        writer = csv.writer(pseudo_buffer)

        def _generate():
            yield '\ufeff'
            if include_headers:
                yield writer.writerow([_sanitize_csv_value(f.name) for f in fields])
            for record in self._iter_records_with_export_data(records, table, fields):
                record_data = read_data(record)
                row = []
                for f in fields:
                    raw_value = self._get_export_field_value(record_data, f, '')
                    if has_system_user_fields and f.field_type in _SYSTEM_USER_FIELD_TYPES:
                        raw_value = _resolve_system_user_value(raw_value, f.field_type, record, user_cache)
                    row.append(self._format_csv_cell(raw_value, f.field_type))
                yield writer.writerow(row)

        return _generate()

    def export_to_excel(
        self,
        table_id: UUID,
        field_ids: Optional[List[UUID]] = None,
        record_ids: Optional[List[UUID]] = None,
        view_id: Optional[UUID] = None,
        include_headers: bool = True,
        sheet_name: str = 'Sheet1',
        rls_context=None,
        view_query: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        """
        导出表格数据为Excel

        Args:
            table_id: 表格ID
            field_ids: 要导出的字段ID列表
            record_ids: 要导出的记录ID列表
            view_id: 视图ID
            include_headers: 是否包含表头
            sheet_name: 工作表名称

        Returns:
            bytes: Excel文件字节内容

        Raises:
            PermissionError: 无查看权限时抛出
            ValueError: 没有可导出的字段时抛出
        """
        # 检查权限
        if not self.check_table_permission(str(table_id), 'viewer'):
            raise PermissionError("没有表格的查看权限")

        table = Table.objects.using(TABDATA_DB_ALIAS).get(id=table_id)

        fields = self._get_export_fields(table_id, field_ids, view_id)
        if not fields:
            raise ValueError("没有可导出的字段")

        records = self._get_records_queryset(
            table_id, record_ids, view_id, max_rows=MAX_EXPORT_ROWS,
            rls_context=rls_context, **(view_query or {}),
        )
        has_system_user_fields = any(f.field_type in _SYSTEM_USER_FIELD_TYPES for f in fields)
        user_cache: Dict[str, str] = {}
        oss_image_cache: Dict[str, Optional[bytes]] = {}

        wb = openpyxl.Workbook(write_only=True)
        safe_sheet_name = re.sub(r'[\\/*?:\[\]]', '_', sheet_name)[:31]
        ws = wb.create_sheet(title=safe_sheet_name)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        left_alignment = Alignment(horizontal="left")
        image_cell_alignment = Alignment(horizontal="left", vertical="top")

        col_max_lengths = [len(field.name) for field in fields]
        for col_idx, field in enumerate(fields, start=1):
            if field.field_type in FILE_BASED_FIELD_TYPES:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = _EXCEL_IMAGE_COL_WIDTH

        if include_headers:
            header_row = []
            for field in fields:
                cell = WriteOnlyCell(ws, value=_sanitize_cell_value(field.name))
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                header_row.append(cell)
            ws.append(header_row)

        row_idx = 1 if include_headers else 0

        for record in self._iter_records_with_export_data(records, table, fields):
            record_data = read_data(record)
            row_cells = []
            row_images = []
            row_height = 0.0
            row_idx += 1
            for col_idx, field in enumerate(fields):
                raw_value = self._get_export_field_value(record_data, field, '')
                if has_system_user_fields and field.field_type in _SYSTEM_USER_FIELD_TYPES:
                    raw_value = _resolve_system_user_value(raw_value, field.field_type, record, user_cache)
                formatted = (
                    _format_excel_date_cell_value(raw_value, field)
                    if field.field_type == 'date'
                    else _format_cell_value(
                        raw_value,
                        field.field_type,
                        bool_style='label',
                    )
                )
                embedded_images = (
                    _build_excel_images_from_file_field(raw_value, oss_image_cache)
                    if field.field_type in FILE_BASED_FIELD_TYPES
                    else []
                )
                has_embedded_images = bool(embedded_images)

                if field.field_type == 'number':
                    try:
                        numeric = float(formatted) if formatted not in ('', None) else None
                        if numeric is not None:
                            raw_digits = formatted.strip().replace('.', '').replace('-', '').replace('+', '')
                            raw_digits = re.sub(r'[eE]\d+$', '', raw_digits)
                            sig_digits = raw_digits.lstrip('0')
                            if len(sig_digits) > 15 or abs(numeric) > 9.99999999999999e15:
                                cell = WriteOnlyCell(ws, value=formatted)
                            else:
                                cell = WriteOnlyCell(ws, value=numeric)
                        else:
                            cell = WriteOnlyCell(ws, value=_sanitize_cell_value(formatted))
                    except (ValueError, TypeError):
                        cell = WriteOnlyCell(ws, value=_sanitize_cell_value(formatted))
                else:
                    cell = WriteOnlyCell(ws, value=_sanitize_cell_value(formatted))
                    if field.field_type == 'date':
                        cell.alignment = left_alignment
                    elif has_embedded_images:
                        cell.alignment = image_cell_alignment

                cell.border = border

                cell_len = len(str(cell.value)) if cell.value else 0
                if cell_len > col_max_lengths[col_idx]:
                    col_max_lengths[col_idx] = cell_len
                if has_embedded_images:
                    col_max_lengths[col_idx] = max(col_max_lengths[col_idx], _EXCEL_IMAGE_COL_WIDTH)
                    y_offset = _EXCEL_IMAGE_TEXT_OFFSET_PX if formatted else 0
                    for embedded_image in embedded_images:
                        row_images.append((col_idx + 1, embedded_image, y_offset))
                        y_offset += int(embedded_image.height) + _EXCEL_IMAGE_STACK_GAP_PX
                    if y_offset > 0:
                        y_offset -= _EXCEL_IMAGE_STACK_GAP_PX
                    row_height = max(
                        row_height,
                        max(_EXCEL_IMAGE_ROW_HEIGHT, y_offset * _EXCEL_POINTS_PER_PIXEL),
                    )

                row_cells.append(cell)
            if row_images:
                try:
                    ws.row_dimensions[row_idx].height = row_height or _EXCEL_IMAGE_ROW_HEIGHT
                except (AttributeError, NotImplementedError):
                    pass
            ws.append(row_cells)

            if row_images:
                for col_number, image, y_offset in row_images:
                    image.anchor = _build_excel_image_anchor(row_idx, col_number, image, y_offset)
                    ws.add_image(image)

        try:
            for col_idx, max_length in enumerate(col_max_lengths, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max(max_length + 2, 10), 50)
        except (AttributeError, NotImplementedError):
            pass

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer.getvalue()

    def export_to_json(
        self,
        table_id: UUID,
        field_ids: Optional[List[UUID]] = None,
        record_ids: Optional[List[UUID]] = None,
        view_id: Optional[UUID] = None,
        format_type: str = 'array',  # 'array' | 'structured' | 'table_full'
        rls_context=None,
        view_query: Optional[Dict[str, Any]] = None,
    ) -> str:
        """
        导出表格数据为JSON（全量加载到内存）

        对于 API 层的 array 格式导出，推荐使用 export_to_json_streaming 以降低内存占用。
        本方法仍保留用于 structured/table_full 格式及非 HTTP 场景。

        Args:
            table_id: 表格ID
            field_ids: 要导出的字段ID列表
            record_ids: 要导出的记录ID列表
            view_id: 视图ID
            format_type: JSON格式类型
                - 'array': [{"field1": "value1", ...}, ...]
                - 'structured': {"headers": [...], "data": [[...], ...]}
                - 'table_full': {"id/name/fields/views/records", ...}

        Returns:
            str: JSON内容

        Raises:
            PermissionError: 无查看权限时抛出
            ValueError: 没有可导出的字段时抛出
        """
        if not self.check_table_permission(str(table_id), 'viewer'):
            raise PermissionError("没有表格的查看权限")

        if format_type == 'table_full':
            table = Table.objects.using(TABDATA_DB_ALIAS).get(id=table_id)

            records = self._get_records_queryset(
                table_id, record_ids, view_id, max_rows=MAX_EXPORT_ROWS,
                rls_context=rls_context, **(view_query or {}),
            )
            result = self._build_table_full_payload(
                table=table,
                records_queryset=records,
            )
            return json.dumps(result, ensure_ascii=False, indent=2)

        fields = self._get_export_fields(table_id, field_ids, view_id)
        if not fields:
            raise ValueError("没有可导出的字段")

        table = Table.objects.using(TABDATA_DB_ALIAS).get(id=table_id)
        records = self._get_records_queryset(
            table_id, record_ids, view_id, max_rows=MAX_EXPORT_ROWS,
            rls_context=rls_context, **(view_query or {}),
        )

        if format_type == 'array':
            data = []
            for record in self._iter_records_with_export_data(records, table, fields):
                record_data = read_data(record)
                row_data = {field.name: self._get_export_field_value(record_data, field, None) for field in fields}
                data.append(row_data)

            return json.dumps(data, ensure_ascii=False, indent=2)

        elif format_type == 'structured':
            headers = [field.name for field in fields]
            data_rows = []

            for record in self._iter_records_with_export_data(records, table, fields):
                record_data = read_data(record)
                row = [self._get_export_field_value(record_data, field, None) for field in fields]
                data_rows.append(row)

            result = {
                'headers': headers,
                'data': data_rows,
                'metadata': {
                    'table_id': str(table_id),
                    'record_count': len(data_rows),
                    'field_count': len(headers),
                    'export_time': timezone.now().isoformat()
                }
            }

            return json.dumps(result, ensure_ascii=False, indent=2)

        else:
            raise ValueError(f"不支持的JSON格式类型: {format_type}")

    def export_to_json_streaming(
        self,
        table_id: UUID,
        field_ids: Optional[List[UUID]] = None,
        record_ids: Optional[List[UUID]] = None,
        view_id: Optional[UUID] = None,
        rls_context=None,
        view_query: Optional[Dict[str, Any]] = None,
    ):
        """
        流式导出 JSON array 格式，返回可迭代的生成器，配合 StreamingHttpResponse 使用。

        仅支持 array 格式（最常用的导出场景），structured 和 table_full 格式
        因结构依赖完整 metadata，仍使用 export_to_json 一次性输出。

        Returns:
            Generator[str]: JSON 片段的生成器

        Raises:
            PermissionError: 无查看权限时抛出
            ValueError: 没有可导出的字段时抛出
        """
        if not self.check_table_permission(str(table_id), 'viewer'):
            raise PermissionError("没有表格的查看权限")

        fields = self._get_export_fields(table_id, field_ids, view_id)
        if not fields:
            raise ValueError("没有可导出的字段")

        table = Table.objects.using(TABDATA_DB_ALIAS).get(id=table_id)
        records = self._get_records_queryset(
            table_id, record_ids, view_id, max_rows=MAX_EXPORT_ROWS,
            rls_context=rls_context, **(view_query or {}),
        )

        def _generate():
            yield '[\n'
            first = True
            for record in self._iter_records_with_export_data(records, table, fields):
                record_data = read_data(record)
                row_data = {field.name: self._get_export_field_value(record_data, field, None) for field in fields}
                row_json = json.dumps(row_data, ensure_ascii=False)
                if not first:
                    yield ',\n  ' + row_json
                else:
                    yield '  ' + row_json
                    first = False
            yield '\n]\n'

        return _generate()

    def export_to_pdf(
        self,
        table_id: UUID,
        field_ids: Optional[List[UUID]] = None,
        record_ids: Optional[List[UUID]] = None,
        view_id: Optional[UUID] = None,
        orientation: str = 'landscape',  # 'portrait' or 'landscape'
        title: Optional[str] = None,
        rls_context=None,
        view_query: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        """
        导出表格数据为PDF

        Args:
            table_id: 表格ID
            field_ids: 要导出的字段ID列表
            record_ids: 要导出的记录ID列表
            view_id: 视图ID
            orientation: 页面方向（portrait竖向/landscape横向）
            title: PDF标题

        Returns:
            bytes: PDF文件字节内容

        Raises:
            PermissionError: 无查看权限时抛出
            ValueError: 没有可导出的字段时抛出
        """
        if not self.check_table_permission(str(table_id), 'viewer'):
            raise PermissionError("没有表格的查看权限")

        table = Table.objects.using(TABDATA_DB_ALIAS).get(id=table_id)

        fields = self._get_export_fields(table_id, field_ids, view_id)
        if not fields:
            raise ValueError("没有可导出的字段")

        records = self._get_records_queryset(
            table_id, record_ids, view_id, max_rows=MAX_EXPORT_ROWS_PDF,
            rls_context=rls_context, **(view_query or {}),
        )
        oss_file_cache: Dict[str, Optional[bytes]] = {}

        font_name = _get_cjk_font_name()

        pdf_buffer = io.BytesIO()
        pagesize = landscape(A4) if orientation == 'landscape' else A4
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=pagesize,
            leftMargin=_PDF_HORIZONTAL_MARGIN,
            rightMargin=_PDF_HORIZONTAL_MARGIN,
            topMargin=_PDF_VERTICAL_MARGIN,
            bottomMargin=_PDF_VERTICAL_MARGIN,
        )

        elements = []
        styles = getSampleStyleSheet()
        title_style = styles['Title'].clone('PDFTitleStyle')
        title_style.keepWithNext = False
        if font_name != 'Helvetica':
            title_style.fontName = font_name
            styles['Normal'].fontName = font_name

        resolved_title = title or table.name or ''

        cell_style = styles['Normal'].clone('CellStyle')
        cell_style.fontSize = 7
        cell_style.leading = 9
        cell_style.wordWrap = 'CJK'
        cell_style.splitLongWords = True
        if font_name != 'Helvetica':
            cell_style.fontName = font_name

        header_style = cell_style.clone('HeaderCellStyle')
        header_style.textColor = colors.whitesmoke
        header_style.fontSize = 8
        header_style.leading = 10

        raw_rows: List[List[Any]] = []
        raw_record_ids: List[str] = []
        user_ids: set[str] = set()
        for record in self._iter_records_with_export_data(records, table, fields):
            record_data = read_data(record)
            raw_row: List[Any] = []
            for field in fields:
                raw_value = self._get_export_field_value(record_data, field, '')
                if field.field_type == 'created_by' and raw_value in (None, ''):
                    raw_value = getattr(record, 'created_by_id', None)
                elif field.field_type == 'last_modified_by' and raw_value in (None, ''):
                    raw_value = getattr(record, 'updated_by_id', None)
                if field.field_type in _USER_FIELD_TYPES:
                    user_ids.update(_collect_user_ids(raw_value))
                raw_row.append(raw_value)
            raw_rows.append(raw_row)
            raw_record_ids.append(str(record.id))

        authorized_files_by_cell = _build_pdf_authorized_files_by_cell(
            table,
            fields,
            raw_record_ids,
        )
        pdf_attachments = _collect_pdf_attachments(
            fields,
            raw_rows,
            raw_record_ids,
            authorized_files_by_cell,
            oss_file_cache,
        )

        user_display_name_by_id, departed_user_ids = _build_pdf_user_display_name_map(
            table.organization_id,
            user_ids,
        )
        formatted_rows: List[List[str]] = []
        for raw_row in raw_rows:
            formatted_row: List[str] = []
            for column_index, field in enumerate(fields):
                raw_value = raw_row[column_index]
                if field.field_type in _USER_FIELD_TYPES:
                    value = _format_pdf_user_value(
                        raw_value,
                        user_display_name_by_id,
                        departed_user_ids,
                    )
                elif field.field_type in _PDF_DATE_FIELD_TYPES:
                    value = _format_excel_date_cell_value(raw_value, field)
                else:
                    value = _format_cell_value(
                        raw_value,
                        field.field_type,
                        bool_style='label',
                        hide_file_image_names=True,
                    )
                if len(value) > PDF_CELL_MAX_LEN:
                    value = value[:PDF_CELL_MAX_LEN - 3] + '...'
                formatted_row.append(value)
            formatted_rows.append(formatted_row)

        view_width_by_field_id: Dict[str, int] = {}
        if view_id:
            view = TableView.objects.get(id=view_id, table_id=table_id)
            column_meta = build_view_column_meta(view, table_fields=fields)
            view_width_by_field_id = {
                field_id: int(meta['width'])
                for field_id, meta in column_meta.items()
                if isinstance(meta, dict)
                and isinstance(meta.get('width'), (int, float))
                and not isinstance(meta.get('width'), bool)
                and meta['width'] > 0
            }

        column_widths = _measure_pdf_column_widths(
            fields,
            formatted_rows,
            font_name,
            view_width_by_field_id,
        )
        column_groups = _partition_pdf_columns(fields, column_widths, doc.width)

        for group_index, column_indexes in enumerate(column_groups):
            if group_index > 0:
                elements.append(PageBreak())
            if resolved_title:
                heading = resolved_title
                if len(column_groups) > 1:
                    heading = f'{resolved_title}（列分片 {group_index + 1}/{len(column_groups)}）'
                elements.append(Paragraph(xml_escape(heading), title_style))

            group_widths = [column_widths[index] for index in column_indexes]
            table_data = [[
                Paragraph(xml_escape(fields[index].name), header_style)
                for index in column_indexes
            ]]
            for row_index, raw_row in enumerate(raw_rows):
                row = []
                for group_column_index, source_column_index in enumerate(column_indexes):
                    field = fields[source_column_index]
                    raw_value = raw_row[source_column_index]
                    value = formatted_rows[row_index][source_column_index]
                    cell_content = Paragraph(xml_escape(value), cell_style)
                    if field.field_type in FILE_BASED_FIELD_TYPES and isinstance(raw_value, (list, dict)):
                        max_image_width = min(
                            _PDF_IMAGE_MAX_WIDTH,
                            max(24.0, group_widths[group_column_index] - _PDF_CELL_HORIZONTAL_PADDING),
                        )
                        image = _build_pdf_image_from_file_field(
                            raw_value,
                            oss_file_cache,
                            authorized_files_by_cell.get(
                                (raw_record_ids[row_index], str(field.id)),
                                {},
                            ),
                            max_image_width,
                        )
                        if image:
                            cell_content = [image, cell_content] if value else image
                    row.append(cell_content)
                table_data.append(row)

            pdf_table = PDFTable(
                table_data,
                colWidths=group_widths,
                repeatRows=1,
                hAlign='LEFT',
            )
            pdf_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ]))
            elements.append(pdf_table)

        canvasmaker = functools.partial(
            _PDFAttachmentCanvas,
            attachments=pdf_attachments,
        )
        doc.build(elements, canvasmaker=canvasmaker)
        pdf_buffer.seek(0)
        return pdf_buffer.getvalue()

    def get_export_stats(
        self,
        table_id: UUID,
        record_ids: Optional[List[UUID]] = None,
        view_id: Optional[UUID] = None,
        view_query: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        """
        获取导出统计信息

        Args:
            table_id: 表格ID
            record_ids: 记录ID列表
            view_id: 视图ID

        Returns:
            Dict: 统计信息

        Raises:
            PermissionError: 无查看权限时抛出
        """
        if not self.check_table_permission(str(table_id), 'viewer'):
            raise PermissionError("没有表格的查看权限")

        table = Table.objects.using(TABDATA_DB_ALIAS).get(id=table_id)
        fields = self._get_export_fields(table_id, view_id=view_id)
        field_count = len(fields)

        records = self._get_records_queryset(table_id, record_ids, view_id, **(view_query or {}))
        record_count = records.count()

        avg_cell_bytes = self._estimate_avg_cell_bytes(table, fields, records)
        total_cells = record_count * field_count
        base_kb = (total_cells * avg_cell_bytes) / 1024.0

        return {
            'table_id': str(table_id),
            'field_count': field_count,
            'record_count': record_count,
            'estimated_size': {
                'csv_kb': round(max(base_kb * 1.1, 1), 1),
                'excel_kb': round(max(base_kb * 1.5 + 10, 5), 1),
                'json_kb': round(max(base_kb * 2.0, 1), 1),
                'pdf_kb': round(max(base_kb * 3.0 + 20, 10), 1),
            }
        }

    def _estimate_avg_cell_bytes(self, table: Table, fields: List[TableField], records_qs, sample_limit: int = 50) -> float:
        """采样部分记录估算平均单元格字节数，避免硬编码常量导致文件大小偏差过大。"""
        sample_fields = fields[:20]
        if not sample_fields:
            return 20.0

        total_bytes = 0
        cell_count = 0
        for record in self._iter_records_with_export_data(records_qs[:sample_limit], table, sample_fields):
            data = read_data(record)
            for field in sample_fields:
                val = self._get_export_field_value(data, field, '')
                if val is None:
                    total_bytes += 0
                elif isinstance(val, str):
                    total_bytes += len(val.encode('utf-8'))
                elif isinstance(val, (list, dict)):
                    total_bytes += len(json.dumps(val, ensure_ascii=False).encode('utf-8'))
                else:
                    total_bytes += len(str(val))
                cell_count += 1

        if cell_count == 0:
            return 20.0
        return max(total_bytes / cell_count, 5.0)
